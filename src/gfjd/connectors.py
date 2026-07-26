"""Controlled source adapters for source-native bronze extraction.

Connectors are declarative TOML records. They may only read files inside the
repository, which means remote material must first pass through the rights-aware
acquisition layer. The adapter preserves source order and labels, adds explicit
GFJD locator columns, and emits a cryptographic receipt for every run.
"""

from __future__ import annotations

import csv
import json
import os
import tomllib
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Literal, cast

from .io import atomic_write_text, canonical_json_bytes, sha256_bytes, sha256_file, write_csv
from .project import Project, load_project

AdapterName = Literal["csv", "json_records", "html_table", "xlsx", "manual_transcription"]


class ConnectorError(RuntimeError):
    """Raised when a connector is unsafe, invalid, or cannot extract its source."""


@dataclass(frozen=True, slots=True)
class ConnectorSpec:
    connector_id: str
    source_id: str
    source_edition_id: str
    adapter: AdapterName
    input_path: str
    output_path: str
    receipt_path: str
    encoding: str = "utf-8-sig"
    delimiter: str = ","
    record_path: str = ""
    table_index: int = 0
    sheet_name: str = ""
    header_row: int = 1
    expected_columns: tuple[str, ...] = ()
    transcription_path: str = ""
    locator_column: str = "_source_locator"
    notes: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class ConnectorResult:
    connector_id: str
    adapter: str
    source_id: str
    source_edition_id: str
    input_path: Path
    output_path: Path
    receipt_path: Path
    rows: int
    columns: tuple[str, ...]
    input_sha256: str
    output_sha256: str
    receipt_sha256: str

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        for field_name in ("input_path", "output_path", "receipt_path"):
            payload[field_name] = str(payload[field_name])
        return payload


@dataclass(frozen=True, slots=True)
class SourceRow:
    source_row_number: int
    source_locator: str
    values: Mapping[str, str]


class _TableParser(HTMLParser):
    """Minimal deterministic HTML table parser with no script execution."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self._table_depth = 0
        self._current_table: list[list[str]] | None = None
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        del attrs
        lower = tag.lower()
        if lower == "table":
            self._table_depth += 1
            if self._table_depth == 1:
                self._current_table = []
        elif lower == "tr" and self._table_depth == 1:
            self._current_row = []
        elif lower in {"th", "td"} and self._table_depth == 1 and self._current_row is not None:
            self._current_cell = []

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        lower = tag.lower()
        if lower in {"th", "td"} and self._current_cell is not None:
            assert self._current_row is not None
            value = " ".join("".join(self._current_cell).split())
            self._current_row.append(value)
            self._current_cell = None
        elif lower == "tr" and self._table_depth == 1 and self._current_row is not None:
            if any(cell != "" for cell in self._current_row):
                assert self._current_table is not None
                self._current_table.append(self._current_row)
            self._current_row = None
        elif lower == "table" and self._table_depth:
            if self._table_depth == 1:
                self.tables.append(self._current_table or [])
                self._current_table = None
            self._table_depth -= 1


def load_connector(project: Project | Path | str, path: Path) -> ConnectorSpec:
    """Load and validate a connector TOML file."""

    resolved_project = _project(project)
    connector_file = _resolve_confined(resolved_project, path)
    try:
        with connector_file.open("rb") as handle:
            payload = tomllib.load(handle)
    except FileNotFoundError as exc:
        raise ConnectorError(f"Connector file does not exist: {connector_file}") from exc
    except tomllib.TOMLDecodeError as exc:
        raise ConnectorError(f"Invalid connector TOML {connector_file}: {exc}") from exc
    raw = payload.get("connector")
    if not isinstance(raw, dict):
        raise ConnectorError("Connector TOML must contain a [connector] table")

    required = {
        "id",
        "source_id",
        "source_edition_id",
        "adapter",
        "input_path",
        "output_path",
        "receipt_path",
    }
    missing = sorted(required - set(raw))
    if missing:
        raise ConnectorError("Connector is missing field(s): " + ", ".join(missing))

    adapter = str(raw["adapter"])
    if adapter not in {
        "csv",
        "json_records",
        "html_table",
        "xlsx",
        "manual_transcription",
    }:
        raise ConnectorError(f"Unsupported adapter {adapter!r}")
    delimiter = str(raw.get("delimiter", ","))
    if len(delimiter) != 1:
        raise ConnectorError("delimiter must be exactly one character")
    header_row = _positive_int(raw.get("header_row", 1), "header_row")
    table_index = _nonnegative_int(raw.get("table_index", 0), "table_index")
    expected_raw = raw.get("expected_columns", [])
    if not isinstance(expected_raw, list) or not all(
        isinstance(value, str) for value in expected_raw
    ):
        raise ConnectorError("expected_columns must be an array of strings")

    spec = ConnectorSpec(
        connector_id=str(raw["id"]).strip(),
        source_id=str(raw["source_id"]).strip(),
        source_edition_id=str(raw["source_edition_id"]).strip(),
        adapter=cast(AdapterName, adapter),
        input_path=str(raw["input_path"]),
        output_path=str(raw["output_path"]),
        receipt_path=str(raw["receipt_path"]),
        encoding=str(raw.get("encoding", "utf-8-sig")),
        delimiter=delimiter,
        record_path=str(raw.get("record_path", "")),
        table_index=table_index,
        sheet_name=str(raw.get("sheet_name", "")),
        header_row=header_row,
        expected_columns=tuple(str(value) for value in expected_raw),
        transcription_path=str(raw.get("transcription_path", "")),
        locator_column=str(raw.get("locator_column", "_source_locator")),
        notes=str(raw.get("notes", "")),
    )
    _validate_spec(resolved_project, spec)
    return spec


def run_connector(
    project: Project | Path | str,
    connector_path: Path,
    *,
    executed_at: datetime | None = None,
) -> ConnectorResult:
    """Run one connector and atomically write bronze output plus its receipt."""

    resolved_project = _project(project)
    spec = load_connector(resolved_project, connector_path)
    input_file = _resolve_confined(resolved_project, Path(spec.input_path))
    output_file = _resolve_confined(resolved_project, Path(spec.output_path))
    receipt_file = _resolve_confined(resolved_project, Path(spec.receipt_path))
    transcription_file = (
        _resolve_confined(resolved_project, Path(spec.transcription_path))
        if spec.transcription_path
        else None
    )
    if not input_file.is_file():
        raise ConnectorError(f"Connector input does not exist: {input_file}")
    _require_output_boundary(resolved_project, output_file, "data/bronze", "build")
    _require_output_boundary(resolved_project, receipt_file, "data/bronze", "build")

    if spec.adapter == "csv":
        headers, rows = _read_csv_rows(input_file, spec)
    elif spec.adapter == "json_records":
        headers, rows = _read_json_rows(input_file, spec)
    elif spec.adapter == "html_table":
        headers, rows = _read_html_rows(input_file, spec)
    elif spec.adapter == "xlsx":
        headers, rows = _read_xlsx_rows(input_file, spec)
    else:
        if transcription_file is None or not transcription_file.is_file():
            raise ConnectorError(f"Manual transcription does not exist: {transcription_file}")
        headers, rows = _read_manual_rows(transcription_file, spec)

    _validate_headers(headers, spec.expected_columns)
    system_headers = (
        "_gfjd_source_id",
        "_gfjd_source_edition_id",
        "_gfjd_connector_id",
        "_gfjd_source_row_number",
        "_gfjd_source_locator",
    )
    collisions = sorted(set(system_headers) & set(headers))
    if collisions:
        raise ConnectorError(
            "Source headers collide with reserved GFJD fields: " + ", ".join(collisions)
        )
    output_headers = [*system_headers, *headers]
    output_rows: list[dict[str, Any]] = []
    for row in rows:
        output_rows.append(
            {
                "_gfjd_source_id": spec.source_id,
                "_gfjd_source_edition_id": spec.source_edition_id,
                "_gfjd_connector_id": spec.connector_id,
                "_gfjd_source_row_number": row.source_row_number,
                "_gfjd_source_locator": row.source_locator,
                **{header: row.values.get(header, "") for header in headers},
            }
        )
    write_csv(output_file, output_headers, output_rows)

    timestamp = _timestamp(executed_at)
    connector_file = _resolve_confined(resolved_project, connector_path)
    receipt_payload: dict[str, Any] = {
        "schema_version": "1.0",
        "connector_id": spec.connector_id,
        "source_id": spec.source_id,
        "source_edition_id": spec.source_edition_id,
        "adapter": spec.adapter,
        "executed_at": timestamp.isoformat().replace("+00:00", "Z"),
        "connector_path": _relative(resolved_project, connector_file),
        "connector_sha256": sha256_file(connector_file),
        "input_path": _relative(resolved_project, input_file),
        "input_sha256": sha256_file(input_file),
        "transcription_path": (
            _relative(resolved_project, transcription_file) if transcription_file else None
        ),
        "transcription_sha256": (sha256_file(transcription_file) if transcription_file else None),
        "output_path": _relative(resolved_project, output_file),
        "output_sha256": sha256_file(output_file),
        "row_count": len(output_rows),
        "source_columns": headers,
        "reserved_columns": list(system_headers),
        "notes": spec.notes,
    }
    receipt_payload["receipt_content_sha256"] = sha256_bytes(canonical_json_bytes(receipt_payload))
    atomic_write_text(
        receipt_file,
        json.dumps(receipt_payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
    )
    return ConnectorResult(
        connector_id=spec.connector_id,
        adapter=spec.adapter,
        source_id=spec.source_id,
        source_edition_id=spec.source_edition_id,
        input_path=input_file,
        output_path=output_file,
        receipt_path=receipt_file,
        rows=len(output_rows),
        columns=tuple(headers),
        input_sha256=receipt_payload["input_sha256"],
        output_sha256=receipt_payload["output_sha256"],
        receipt_sha256=sha256_file(receipt_file),
    )


def verify_connector_receipt(
    project: Project | Path | str,
    receipt_path: Path,
) -> list[str]:
    """Verify receipt integrity and the current input/output objects."""

    resolved_project = _project(project)
    path = _resolve_confined(resolved_project, receipt_path)
    errors: list[str] = []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return [f"Could not read connector receipt: {exc}"]
    if not isinstance(payload, dict):
        return ["Connector receipt root must be an object"]
    if payload.get("schema_version") != "1.0":
        errors.append("Unsupported connector receipt schema_version")
    content_hash = str(payload.get("receipt_content_sha256") or "")
    without_hash = dict(payload)
    without_hash.pop("receipt_content_sha256", None)
    expected_content_hash = sha256_bytes(canonical_json_bytes(without_hash))
    if content_hash != expected_content_hash:
        errors.append("Connector receipt content hash mismatch")
    for label in ("connector", "input", "output"):
        relative = str(payload.get(f"{label}_path") or "")
        digest = str(payload.get(f"{label}_sha256") or "")
        if not relative:
            errors.append(f"Receipt has no {label}_path")
            continue
        try:
            candidate = _resolve_confined(resolved_project, Path(relative))
        except ConnectorError as exc:
            errors.append(str(exc))
            continue
        if not candidate.is_file():
            errors.append(f"Receipt {label} file is missing: {relative}")
        elif sha256_file(candidate) != digest:
            errors.append(f"Receipt {label} checksum mismatch: {relative}")
    transcription_relative = payload.get("transcription_path")
    if transcription_relative:
        transcription_digest = str(payload.get("transcription_sha256") or "")
        try:
            transcription = _resolve_confined(resolved_project, Path(str(transcription_relative)))
        except ConnectorError as exc:
            errors.append(str(exc))
        else:
            if not transcription.is_file():
                errors.append(f"Receipt transcription file is missing: {transcription_relative}")
            elif sha256_file(transcription) != transcription_digest:
                errors.append(f"Receipt transcription checksum mismatch: {transcription_relative}")
    return errors


def _project(value: Project | Path | str) -> Project:
    return value if isinstance(value, Project) else load_project(Path(value))


def _resolve_confined(project: Project, path: Path) -> Path:
    candidate = path.expanduser()
    resolved = (
        candidate.resolve() if candidate.is_absolute() else (project.root / candidate).resolve()
    )
    try:
        resolved.relative_to(project.root)
    except ValueError as exc:
        raise ConnectorError(f"Connector path escapes repository root: {path}") from exc
    return resolved


def _require_output_boundary(project: Project, path: Path, *allowed_roots: str) -> None:
    for root_name in allowed_roots:
        try:
            path.relative_to(project.root / root_name)
            return
        except ValueError:
            continue
    raise ConnectorError(
        f"Connector output must be under one of {', '.join(allowed_roots)}: {path}"
    )


def _validate_spec(project: Project, spec: ConnectorSpec) -> None:
    del project
    if not spec.connector_id or not spec.source_id or not spec.source_edition_id:
        raise ConnectorError("connector id, source_id, and source_edition_id must be non-empty")
    if any(char.isspace() for char in spec.connector_id):
        raise ConnectorError("connector id must not contain whitespace")
    if spec.adapter == "manual_transcription":
        if not spec.transcription_path:
            raise ConnectorError("manual_transcription requires transcription_path")
        if not spec.locator_column.strip():
            raise ConnectorError("manual_transcription requires locator_column")


def _read_csv_rows(path: Path, spec: ConnectorSpec) -> tuple[list[str], list[SourceRow]]:
    try:
        with path.open("r", encoding=spec.encoding, newline="") as handle:
            reader = csv.reader(handle, delimiter=spec.delimiter)
            all_rows = list(reader)
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ConnectorError(f"Could not read CSV source {path}: {exc}") from exc
    return _matrix_to_rows(all_rows, spec.header_row, locator_prefix="row")


def _read_json_rows(path: Path, spec: ConnectorSpec) -> tuple[list[str], list[SourceRow]]:
    try:
        payload = json.loads(path.read_text(encoding=spec.encoding))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ConnectorError(f"Could not read JSON source {path}: {exc}") from exc
    selected: Any = payload
    if spec.record_path:
        for token in spec.record_path.split("."):
            if isinstance(selected, dict) and token in selected:
                selected = selected[token]
            else:
                raise ConnectorError(f"record_path token {token!r} was not found")
    if not isinstance(selected, list):
        raise ConnectorError("JSON record_path must resolve to an array")
    records: list[dict[str, Any]] = []
    headers: list[str] = []
    for index, item in enumerate(selected):
        if not isinstance(item, dict):
            raise ConnectorError(f"JSON record {index} is not an object")
        record = {str(key): value for key, value in item.items()}
        records.append(record)
        for key in record:
            if key not in headers:
                headers.append(key)
    _validate_unique_headers(headers)
    rows = [
        SourceRow(
            source_row_number=index + 1,
            source_locator=f"{spec.record_path or '$'}[{index}]",
            values={header: _normalise_value(record.get(header)) for header in headers},
        )
        for index, record in enumerate(records)
    ]
    return headers, rows


def _read_html_rows(path: Path, spec: ConnectorSpec) -> tuple[list[str], list[SourceRow]]:
    try:
        text = path.read_text(encoding=spec.encoding)
    except (OSError, UnicodeError) as exc:
        raise ConnectorError(f"Could not read HTML source {path}: {exc}") from exc
    parser = _TableParser()
    parser.feed(text)
    if spec.table_index >= len(parser.tables):
        raise ConnectorError(
            f"HTML table_index {spec.table_index} not found; "
            f"source has {len(parser.tables)} table(s)"
        )
    return _matrix_to_rows(
        parser.tables[spec.table_index],
        spec.header_row,
        locator_prefix=f"table:{spec.table_index}/row",
    )


def _read_manual_rows(path: Path, spec: ConnectorSpec) -> tuple[list[str], list[SourceRow]]:
    headers, rows = _read_csv_rows(path, spec)
    if spec.locator_column not in headers:
        raise ConnectorError(
            f"Manual transcription is missing locator column {spec.locator_column!r}"
        )
    data_headers = [header for header in headers if header != spec.locator_column]
    if not data_headers:
        raise ConnectorError("Manual transcription has no data columns")
    manual_rows: list[SourceRow] = []
    for row in rows:
        locator = row.values.get(spec.locator_column, "").strip()
        if not locator:
            raise ConnectorError(
                f"Manual transcription row {row.source_row_number} has no provenance locator"
            )
        manual_rows.append(
            SourceRow(
                source_row_number=row.source_row_number,
                source_locator=locator,
                values={header: row.values.get(header, "") for header in data_headers},
            )
        )
    return data_headers, manual_rows


def _read_xlsx_rows(path: Path, spec: ConnectorSpec) -> tuple[list[str], list[SourceRow]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise ConnectorError("XLSX adapter requires the optional openpyxl package") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[spec.sheet_name] if spec.sheet_name else workbook[workbook.sheetnames[0]]
        matrix = [
            [_normalise_value(value) for value in row] for row in sheet.iter_rows(values_only=True)
        ]
        title = sheet.title
        workbook.close()
    except (OSError, KeyError, ValueError) as exc:
        raise ConnectorError(f"Could not read XLSX source {path}: {exc}") from exc
    headers, rows = _matrix_to_rows(matrix, spec.header_row, locator_prefix=f"{title}!row")
    return headers, rows


def _matrix_to_rows(
    matrix: Sequence[Sequence[Any]],
    header_row: int,
    *,
    locator_prefix: str,
) -> tuple[list[str], list[SourceRow]]:
    if len(matrix) < header_row:
        raise ConnectorError(f"Source has no configured header row {header_row}")
    headers = [_normalise_header(value) for value in matrix[header_row - 1]]
    while headers and headers[-1] == "":
        headers.pop()
    if not headers:
        raise ConnectorError("Source header row is empty")
    if any(header == "" for header in headers):
        raise ConnectorError("Source header row contains a blank column name")
    _validate_unique_headers(headers)
    rows: list[SourceRow] = []
    for matrix_index, raw_row in enumerate(matrix[header_row:], start=header_row + 1):
        values = list(raw_row[: len(headers)])
        values.extend([""] * (len(headers) - len(values)))
        if not any(_normalise_value(value) != "" for value in values):
            continue
        rows.append(
            SourceRow(
                source_row_number=matrix_index,
                source_locator=f"{locator_prefix}:{matrix_index}",
                values={
                    header: _normalise_value(values[index]) for index, header in enumerate(headers)
                },
            )
        )
    return headers, rows


def _normalise_header(value: Any) -> str:
    return " ".join(_normalise_value(value).split())


def _normalise_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return str(value)


def _validate_unique_headers(headers: Iterable[str]) -> None:
    values = list(headers)
    duplicates = sorted({value for value in values if values.count(value) > 1})
    if duplicates:
        raise ConnectorError("Source contains duplicate headers: " + ", ".join(duplicates))


def _validate_headers(headers: Sequence[str], expected: Sequence[str]) -> None:
    _validate_unique_headers(headers)
    if expected and list(headers) != list(expected):
        raise ConnectorError(
            "Source columns do not match connector contract; expected "
            f"{list(expected)!r}, found {list(headers)!r}"
        )


def _positive_int(value: Any, name: str) -> int:
    result = _nonnegative_int(value, name)
    if result < 1:
        raise ConnectorError(f"{name} must be at least 1")
    return result


def _nonnegative_int(value: Any, name: str) -> int:
    if isinstance(value, bool):
        raise ConnectorError(f"{name} must be an integer")
    try:
        result = int(value)
    except (TypeError, ValueError) as exc:
        raise ConnectorError(f"{name} must be an integer") from exc
    if result < 0:
        raise ConnectorError(f"{name} must be non-negative")
    return result


def _timestamp(value: datetime | None) -> datetime:
    if value is not None:
        if value.tzinfo is None:
            raise ConnectorError("executed_at must include a timezone")
        return value.astimezone(UTC).replace(microsecond=0)
    source_epoch = os.getenv("SOURCE_DATE_EPOCH")
    if source_epoch:
        return datetime.fromtimestamp(int(source_epoch), UTC).replace(microsecond=0)
    return datetime.now(UTC).replace(microsecond=0)


def _relative(project: Project, path: Path) -> str:
    return path.relative_to(project.root).as_posix()
